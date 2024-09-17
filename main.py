# -*- coding: utf-8 -*-
"""
Created on Mon Sep 16 09:30:00 2024
"""

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import pickle
import csv
import seaborn as sns
from matplotlib.patches import Patch
from pathlib import Path
import os
from api import fetch_user_permissions

# Funktion für Ähnlichkeitsberechnung
def similarity(A):
    number_rows_A, number_columns_A = A.shape
    a = np.dot(A, A.transpose())
    A_ = np.logical_not(A).astype(int)
    d = np.dot(A_, A_.transpose())
    return (a + d) / number_columns_A

# Finden des Elements in der Matrix, das die maximale Zahl hat (für 2 Benutzer mit den ähnlichsten Rollen)
def find_max(A):
    max_value = None
    max_row = None
    max_col = None
    for i in range(len(A)):
        for j in range(i+1, len(A[i])):
            if max_value is None or A[i, j] > max_value:
                max_value = A[i, j]
                max_row = i
                max_col = j
    return (max_row, max_col)

# Darstellen welche Benutzer die neue Rolle erhalten
def update_A(A, i, j):
    A[i] = A[i] * A[j]
    A = np.delete(A, j, axis=0)
    return A

# Erstellen der neuen Rolle durch Verbindung der ähnlichsten Rollen
def update_C(C, i, j):
    C[i] = C[i] + C[j]
    C = np.delete(C, j, axis=0)
    return C

# Berechnung der Benutzeranzahl pro Rolle
def update_alpha(A):
    _, N = A.shape
    return np.dot(A, np.ones((N, 1)))

# Berechnung der Anzahl der Berechtigungen pro Rolle
def update_x(C):
    _, M = C.shape
    return np.dot(C, np.ones((M, 1)))

# Berechnung des Beitrags jeder Rolle
def update_y(alpha, x):
    return alpha * x

# Berechnung der Summe der Beiträge der k größten Rollen
def largest_roles_(y, k):
    summ = 0
    largest_indices = np.argsort(y, axis=0)[-k:]
    for i in largest_indices:
        summ += y[i]
    return largest_indices, summ

# Prüfen, ob ein Benutzer alle Berechtigungen einer Rolle hat
def has_all_rights(user, rights, matrix, list_of_users, list_of_permissions):
    user_index = list_of_users.index(user)
    right_indices = [list_of_permissions.index(right) for right in rights]
    return all(matrix[right, user_index].astype(int) == 1 for right in right_indices)

# Zeilen in einer Matrix verschieben
def verschiebe_zeilen(array, index_list):
    flatten_indices = [index for sublist in index_list for index in sublist]
    remaining_indices = [i for i in range(array.shape[0]) if i not in flatten_indices]
    new_order = flatten_indices + remaining_indices
    return array[new_order,:]

def verschiebe_zeilen_2(matrix, index_list):
    flatten_indices = [index for sublist in index_list for index in sublist]
    remaining_indices = [i for i in range(matrix.shape[0]) if i not in flatten_indices]
    new_order = flatten_indices + remaining_indices
    return matrix[new_order, :]

def verschiebe_zeilen_3(matrix, index_list):
    matrix = list(matrix)
    flatten_indices = [index for sublist in index_list for index in sublist]
    remaining_indices = [i for i in range(len(matrix)) if i not in flatten_indices]
    new_order = flatten_indices + remaining_indices
    return [matrix[i] for i in new_order]


#    except ValueError as e:
#        print(f"IndexError: {e}")
#        return False

# Erstellen eines Pfads relativ zum Arbeitsverzeichnis
base_path = Path(__file__).parent

# Gruppenname für den Aufruf
gruppenname = input("Enter den Gruppenname für den Aufruf: ")

## Aufruf der Funktion mit dem Gruppennamen und dem Pfad, in dem die Ergebnisse gespeichert werden sollen
fetch_user_permissions(base_path, gruppenname)

# Definieren der relativen Pfade
user_permissions_path = base_path / "user_permissions.csv"
permissions_list_path = base_path / "data_permissions.csv"
list_of_users_path = base_path / "list_of_users.txt"

df = pd.read_csv(user_permissions_path, encoding="latin1", header=None)
ups_erweiterte = df.drop(df.columns[0], axis=1)
up = ups_erweiterte.to_numpy()
up = up.transpose()
# Initialisation
k_max = int(input("Enter die Anzahl von den Rollen: "))
# Maximale Anzahl von den Rollen (welche Anzahl von Rollen nehmen wir, um den Nutzenfunktion zu berechnen)

folder_path = input("Geben den Pfad, wo die Ergebnisse gespeichert werden sollen: ")

# Anzahl von Berechtigungen und Benutzer einlesen
number_permissions, number_users = up.shape

# Maximal mögliche Deckung (der maximalen Wert von Nutzenfunktion)
total_number_ups = np.sum(up)

# Anfangswert für A
roles_to_users_a = up
print(np.shape(roles_to_users_a))

# Anfangswert für C
roles_to_permissions_c = np.eye(number_permissions)

# Listen laden
# Liste von den Berechtigungen
# with open("list_of_permissions.txt", "rb") as file:
#    list_of_permissions = pickle.load(file)
#    print(f"number of permissions : {len(list_of_permissions)}")
df = pd.read_csv(permissions_list_path, header=None, encoding='latin1')

# Umwandeln des DataFrames in eine Liste (jede Zeile wird ein Element)
list_of_permissions = df.values.flatten().tolist()

# Ausgabe der Liste
print(len(list_of_permissions))

# Liste von den Benutzern
with open("list_of_users.txt", "rb") as file:
    list_of_users = pickle.load(file)

# Maximale Anzahl von den Iterationen
print(number_permissions)
k = number_permissions
print(k)
# die Liste für k und entsprechende Deckung erstellen
shares = []
ks = []
share_optimum = 0
k_optimum = 0
# Iterationen
while (k > 1):
    # Ähnlichkeitsberechnung
    similarity_matrix = similarity(roles_to_users_a)

    # Nummer von Rollen, die "am gleichsten" sind
    i, j = find_max(similarity_matrix)

    # Aktualisierung von A
    roles_to_users_a = update_A(roles_to_users_a, i, j)

    # Aktualisierung von C
    roles_to_permissions_c = update_C(roles_to_permissions_c, i, j)
    roles_to_permissions_c = roles_to_permissions_c.astype(int)

    # Aktualisierung von Alpha
    number_users_assigned_candidate_role_alpha = update_alpha(roles_to_users_a)

    # Aktualisierung von x
    number_permissions_assigned_to_candidate_role_x = update_x(roles_to_permissions_c)

    # Berechnung von den Beitrag für jede Rolle
    cost_function_y = update_y(number_users_assigned_candidate_role_alpha,
                               number_permissions_assigned_to_candidate_role_x)

    # Falls die Anzahl von Rollen weniger als k_max, dann machen wir k_max gleich k
    if k_max > k:
        k_func = k
    else:
        k_func = k_max

    # Rollen, die maximale Deckung haben, herausfinden und die Deckung für die berechnen
    largest_roles_indices, Y = largest_roles_(cost_function_y, k_func)

    # maximale und aktualle Deckung vergleichen (1 ist das beste Ergebnis)
    share = Y / total_number_ups

    # die Werte von k und entsprechende Deckung in Liste hinzufügen
    shares.append(share[0])
    ks.append(k)
    if (share >= share_optimum):
        share_optimum = share
        k_optimum = k
        largest_roles_indices_optimum = largest_roles_indices
        roles_to_permissions_c_optimum = roles_to_permissions_c
        roles_to_users_a_optimum = roles_to_users_a

    print(f" K_optimum = {k_optimum}")
    print(f" Share_optimum = {share_optimum}")

    # die Werte von k und entsprechende Deckung betrachten und solche k wählen, die maximale Deckung versorgen
    print(f"k={k}")
    print(share)
    k -= 1

    if (k == 1):

        #        permission_index_für_alle_rollen = []
        #        list_of_names_of_permissions_for_roles = []
        #
        #        for role_index in largest_roles_indices:  # only consider the roles in the largest k roles
        #            list_of_names_of_permissions_for_each_role = []  # list of permissions for each individual role
        #
        #    # Only take the permissions that correspond to 1
        #        permissions_indices_for_each_role = np.where(roles_to_permissions_c[role_index] == 1)
        #
        #    # Access the correct indices based on np.where result structure
        #        if len(permissions_indices_for_each_role) > 1:
        #            permission_indices = permissions_indices_for_each_role[1]
        #        else:
        #            permission_indices = permissions_indices_for_each_role[0]
        #
        #        permission_index_für_alle_rollen.append(permission_indices)
        #
        #        for permission_index in permission_indices:
        #            if 0 <= permission_index < len(list_of_permissions):
        #                list_of_names_of_permissions_for_each_role.append(list_of_permissions[permission_index])
        #            else:
        #                print(f"Invalid permission_index: {permission_index}, skipping...")
        #
        #    # Add the list of permission names for this role to the final list
        #        list_of_names_of_permissions_for_roles.append(list_of_names_of_permissions_for_each_role)

        permission_index_für_alle_rollen = []
        for role_index in largest_roles_indices_optimum:  # nur solche Rolle betrachten, die in k maximalen Rollen sind
            permissions_indices_for_each_role = np.where(roles_to_permissions_c_optimum[role_index] == 1)
            permission_index_für_alle_rollen.append(permissions_indices_for_each_role[1])

        list_of_names_of_permissions_for_roles = []  # eine Liste von Namen von den Berechtigungen für jede Rolle
        for role_index in largest_roles_indices_optimum:  # nur solche Rolle betrachten, die in k maximalen Rollen sind
            list_of_names_of_permissions_for_each_role = []  # Liste von den Berechtigungen für jede einzelne Rolle
            # Nur solche Berechtigungen in die Liste aufnehmen, die 1 entsprechen
            permissions_indices_for_each_role = np.where(roles_to_permissions_c_optimum[role_index] == 1)
            for permission_index in permissions_indices_for_each_role[1]:
                list_of_names_of_permissions_for_each_role.append(list_of_permissions[permission_index])
            # Hinzufügen eine Zeile von jeder Rolle
            list_of_names_of_permissions_for_roles.append(list_of_names_of_permissions_for_each_role)

        #        up1 = [[1, 0, 0],[0, 1, 0],[0,0,1]]
        #        up2 = np.array(up1)
        #        p_i = [[2, 1]]
        #        up1 = verschiebe_zeilen(up2, p_i)
        #        print(up1)

        up = verschiebe_zeilen(up, permission_index_für_alle_rollen)
        roles_to_permissions_c_optimum = verschiebe_zeilen(roles_to_permissions_c_optimum.transpose(),
                                                           permission_index_für_alle_rollen).transpose()
        roles_to_users_a_optimum = verschiebe_zeilen_2(roles_to_users_a_optimum, largest_roles_indices_optimum)
        list_of_permissions = verschiebe_zeilen_3(np.array(list_of_permissions), permission_index_für_alle_rollen)
        #        print(len(list_of_names_of_permissions_for))
        #        print(len(list_of_permissions))
        # print(len(np.array(list_of_permissions)))
        # print(permission_index_für_alle_rollen)
        largest_roles_indices_optimum = [[i] for i in range(k_max)]

        # Aktualisierung von permission_index_für_alle_rollen
        new_number_of_permission = 0
        permission_index_für_alle_rollen_new = []
        for index_role in range(len(permission_index_für_alle_rollen)):
            permissions_indices_for_each_role_new = []
            for index_perm in range(len(permission_index_für_alle_rollen[index_role])):
                permissions_indices_for_each_role_new.append(new_number_of_permission)
                new_number_of_permission += 1
            permission_index_für_alle_rollen_new.append(permissions_indices_for_each_role_new)

        #        permission_index_für_alle_rollen = []
        #        list_of_names_of_permissions_for_roles = []
        #        for role_index in largest_roles_indices: # nur solche Rolle betrachten, die in k maximalen Rollen sind
        #            list_of_names_of_permissions_for_each_role = [] # Liste von den Berechtigungen für jede einzelne Rolle
        #            # Nur solche Berechtigungen in die Liste aufnehmen, die 1 entsprechen
        #            permissions_indices_for_each_role = np.where(roles_to_permissions_c[role_index] == 1)
        #            permission_index_für_alle_rollen.append(permissions_indices_for_each_role[1])
        #            for permission_index in permissions_indices_for_each_role[1]:
        #                list_of_names_of_permissions_for_each_role.append(list_of_permissions[permission_index])
        #            # Hinzufügen eine Zeile von jed        list_of_names_of_permissions_for_roles = [] # eine Liste von Namen von den Berechtigungen für jede Rolle
        #            list_of_names_of_permissions_for_roles.append(list_of_names_of_permissions_for_each_role)

        # das Ergebnis notiren (aie Liste von den Berechtigungen für jede Rolle)

        with open(permissions_list_path, "w", newline="") as csvfile:
            writer = csv.writer(csvfile)
            for each_role in range(len(list_of_names_of_permissions_for_roles)):
                writer.writerow([list_of_names_of_permissions_for_roles[each_role]])
        print(largest_roles_indices_optimum)

        users_index_für_alle_rollen = []
        for role_index in largest_roles_indices_optimum:  # nur solche Rolle betrachten, die in k maximalen Rollen sind
            users_indices_for_each_role = np.where(roles_to_users_a_optimum[role_index] == 1)
            users_index_für_alle_rollen.append(users_indices_for_each_role[1])

        print(users_index_für_alle_rollen)

        list_of_names_of_users_for_roles = []  # eine Liste von Benutzern für jede Rolle
        for role_index in largest_roles_indices_optimum:  # nur solche Rolle betrachten, die in k maximalen Rollen sind
            list_of_names_of_users_for_each_role = []  # Liste von den Benutzern für jede einzelne Rolle
            users_indices_for_each_role = np.where(roles_to_users_a_optimum[role_index] == 1)
            for user_index in users_indices_for_each_role[1]:
                list_of_names_of_users_for_each_role.append(list_of_users[user_index])
            # Hinzufügen eine Zeile von jeder Rolle
            list_of_names_of_users_for_roles.append(list_of_names_of_users_for_each_role)

        # das Ergebnis notiren (aie Liste von den Benutzern für jede Rolle)
        csv_filename = "users_to_roles.csv"
        with open(csv_filename, "w", newline="") as csvfile:
            writer = csv.writer(csvfile)
            for each_role in range(len(list_of_names_of_users_for_roles)):
                writer.writerow([list_of_names_of_users_for_roles[each_role]])

        # Farben für die Rollen
        cmap = plt.cm.get_cmap("tab20", k_max)
        role_colors = [cmap(i) for i in range(k_max)]

        # Dictionary für Berechtigungen und ihre Farben
        right_colors = {}
        for idx, rights in enumerate(list_of_names_of_permissions_for_roles):
            for right in rights:
                right_colors[right] = role_colors[idx]

        colors = np.empty_like(up, dtype=object)
        # colors = np.full_like(up, "gray", dtype=object)
        # colors = np.where(up == 1, "gray", "")

        binary_result = np.zeros((number_permissions, number_users), dtype=int)

        for i in range(number_permissions):
            for j in range(number_users):
                if (up[i, j].astype(int) == 1):
                    colors[i, j] = (0.5, 0.5, 0.5, 1.0)
                    binary_result[i, j] = -1

        for role_idx, rights in enumerate(list_of_names_of_permissions_for_roles):
            for user in list_of_names_of_users_for_roles[role_idx]:
                if has_all_rights(user, rights, up, list_of_users, list_of_permissions):
                    for right in rights:
                        right_idx = list_of_permissions.index(right)
                        user_idx = list_of_users.index(user)
                        colors[right_idx, user_idx] = role_colors[role_idx]
                        binary_result[right_idx, user_idx] = role_idx + 1
                        if (role_idx == 1):
                            print(f"Assigned color {role_colors[role_idx]} to cell ({right_idx}, {user_idx})")

        for role in range(k_max):
            for user in range(number_users):
                if user not in users_index_für_alle_rollen[role]:
                    k_0 = 0
                    k_1 = 0
                    for permission in permission_index_für_alle_rollen_new[role]:
                        if up[permission][user] == 1:
                            k_1 += 1
                        else:
                            k_0 += 1
                    if (k_1 / (k_1 + k_0) >= 0.7):
                        users_index_für_alle_rollen[role] = np.append(users_index_für_alle_rollen[role], user)
                        print(share_optimum)
                        share_optimum += k_1 / total_number_ups
                        print(share_optimum)
                        for permission in permission_index_für_alle_rollen_new[role]:
                            if up[permission][user] == 1:
                                colors[permission, user] = role_colors[role]
                                binary_result[permission, user] = role + 1
                            elif up[permission][user] == 0:
                                colors[permission, user] = (1, 0, 0)
                                binary_result[permission, user] = 100

        # Visualize the heatmap with colors assigned
        fig, ax = plt.subplots(figsize=(40, 120))
        sns.heatmap(up, annot=True, fmt="d", cmap="coolwarm", cbar=False, ax=ax, linewidths=.1, linecolor="black")

        # Add colored rectangles for roles
        for i in range(up.shape[0]):
            for j in range(up.shape[1]):
                if isinstance(colors[i, j], tuple):
                    rect = plt.Rectangle((j, i), 1, 1, fill=False, edgecolor=colors[i, j], linewidth=3)
                    ax.add_patch(rect)
                else:
                    rect = plt.Rectangle((j, i), 1, 1, fill=False, edgecolor="white", linewidth=3)
                    ax.add_patch(rect)

        column_labels = list_of_users

        ax.set_xticks(np.arange(len(column_labels)) + 1)
        ax.set_xticklabels(column_labels, rotation=45, ha='right')
        ax.xaxis.tick_top()  # X-Achsen-Ticks oben platzieren

        row_labels = list_of_permissions
        ax.set_yticks(np.arange(len(row_labels)) + 0.5)
        ax.set_yticklabels(row_labels, rotation=0, ha='right', fontsize=6)

        ax.set_ylim(len(row_labels), 0)
        plt.tight_layout()

        legend_elements = [Patch(facecolor=role_colors[idx], edgecolor='r', label=f'Role {idx + 1}') for idx in
                           range(k_max)]
        fig.legend(handles=legend_elements, loc='upper right', bbox_to_anchor=(0.5, -0.05), ncol=3)
        fig.tight_layout()

        # Dateipfad erstellen
        file_path = os.path.join(folder_path, 'heatmap1.png')

        plt.savefig(file_path)
        plt.show()

        row_names = [f'{permission}' for permission in (list_of_permissions)]
        column_names = [f'{user}' for user in (list_of_users)]

        # Umwandlung der Matrix in ein DataFrame
        df = pd.DataFrame(binary_result, index=row_names, columns=column_names)

        # Fester Dateiname
        file_name = 'matrix.xlsx'

        # Vollständiger Pfad zur Datei
        full_path = os.path.join(folder_path, file_name)

        # DataFrame in Excel-Datei speichern
        df.to_excel(full_path, index=True, header=True)

        # Eintrag in Zelle A1 setzen
        from openpyxl import load_workbook

        # Lade die Excel-Datei
        wb = load_workbook(full_path)
        ws = wb.active

        # Setze den Eintrag in Zelle A1
        ws['A1'] = 'Benutzer/AD-Gruppe'

        # Speichere die Änderungen
        wb.save(full_path)
        print(f"Ergebnisse wurden erfolgreich in {full_path} gespeichert.")

# Grafik: share(k). Auf solche Weise können wir solche k herausfinden, die maximale Deckung versorgen
plt.figure(figsize=(10, 6))
plt.plot(ks, shares, marker="o")
plt.xlabel("Number of roles (M)")
plt.ylabel("Share")
plt.title("Share vs Number of roles (M)")
plt.grid(True)
plt.show()





